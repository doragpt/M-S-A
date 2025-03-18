
import pandas as pd
from datetime import datetime
import pytz
import os
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, Color
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.spreadsheet_drawing import AnchorMarker, OneCellAnchor
from openpyxl.utils import units
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.formatting.rule import ColorScaleRule

class ReportGenerator:
    def __init__(self):
        # モダンなカラーパレット
        self.header_fill = PatternFill(start_color="1A73E8", end_color="1A73E8", fill_type="solid")
        self.header_font = Font(color="FFFFFF", bold=True, size=11, name='Yu Gothic')
        self.title_font = Font(color="202124", bold=True, size=14, name='Yu Gothic')
        self.border = Border(
            left=Side(style='thin', color="E8EAED"),
            right=Side(style='thin', color="E8EAED"),
            top=Side(style='thin', color="E8EAED"),
            bottom=Side(style='thin', color="E8EAED")
        )

    def generate_all_stores_report(self, stores_data, output_path):
        """全店舗の詳細Excelレポートを生成"""
        if not stores_data:
            raise ValueError("店舗データが空です")

        output_path = output_path.replace('.pdf', '.xlsx')
        
        try:
            # 期間情報の取得
            dates = [store.get('timestamp') for store in stores_data if store.get('timestamp')]
            start_date = min(dates).strftime('%Y/%m/%d %H:%M') if dates else 'N/A'
            end_date = max(dates).strftime('%Y/%m/%d %H:%M') if dates else 'N/A'
            
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                # サマリーシート
                total_stores = len(stores_data)
                total_working = sum(store.get('working_staff', 0) for store in stores_data)
                total_active = sum(store.get('active_staff', 0) for store in stores_data)
                avg_rate = sum(store.get('rate', 0) for store in stores_data) / total_stores if total_stores > 0 else 0

                summary_data = {
                    '項目': ['集計期間（開始）', '集計期間（終了）', '総店舗数', '総勤務人数', '総即ヒメ数', '平均稼働率'],
                    '値': [
                        start_date,
                        end_date,
                        total_stores,
                        total_working,
                        total_active,
                        f"{avg_rate:.1f}%"
                    ]
                }
                total_working = sum(store.get('working_staff', 0) for store in stores_data)
                total_active = sum(store.get('active_staff', 0) for store in stores_data)
                avg_rate = sum(store.get('rate', 0) for store in stores_data) / total_stores if total_stores > 0 else 0

                summary_data = {
                    '項目': ['総店舗数', '総勤務人数', '総即ヒメ数', '平均稼働率'],
                    '値': [
                        total_stores,
                        total_working,
                        total_active,
                        f"{avg_rate:.1f}%"
                    ]
                }
                summary_df = pd.DataFrame(summary_data)
                summary_df.to_excel(writer, sheet_name='サマリー', index=False)
                
                # サマリーシートのスタイル設定
                ws = writer.sheets['サマリー']
                self._apply_sheet_styling(ws, summary_df)
                
                # 店舗詳細シート
                store_details = []
                for store in stores_data:
                    store_details.append({
                        '店舗名': store.get('store_name', ''),
                        '業種': store.get('biz_type', ''),
                        'ジャンル': store.get('genre', ''),
                        'エリア': store.get('area', ''),
                        '稼働率': store.get('rate', 0),
                        '勤務人数': store.get('working_staff', 0),
                        '即ヒメ数': store.get('active_staff', 0)
                    })
                
                details_df = pd.DataFrame(store_details)
                details_df.to_excel(writer, sheet_name='店舗詳細', index=False)
                
                # 店舗詳細シートのスタイル設定
                ws = writer.sheets['店舗詳細']
                self._apply_sheet_styling(ws, details_df)
                
                # 稼働率の条件付き書式
                self._apply_conditional_formatting(ws, details_df)
                
                # エリア分析シートの追加
                self._create_area_analysis_sheet(writer, store_details)
                
                # 時間帯別分析シートの追加
                self._create_time_analysis_sheet(writer, stores_data)
                
                # ジャンル分析シートの追加
                self._create_genre_analysis_sheet(writer, store_details)
                
                # 全シートの幅調整と体裁整理
                for sheet_name in writer.sheets:
                    ws = writer.sheets[sheet_name]
                    self._adjust_column_widths(ws)
                    self._apply_sheet_styling(ws, pd.DataFrame())  # 基本スタイルを適用
                
            return output_path

    def _create_time_analysis_sheet(self, writer, stores_data):
        """時間帯別分析シートを作成"""
        if not stores_data:
            return
            
        # 時間帯別にデータを集計
        time_stats = {}
        for store in stores_data:
            timestamp = store.get('timestamp')
            if not timestamp:
                continue
                
            hour = timestamp.hour
            if hour not in time_stats:
                time_stats[hour] = {
                    'store_count': 0,
                    'total_rate': 0,
                    'working_staff': 0,
                    'active_staff': 0
                }
            
            time_stats[hour]['store_count'] += 1
            time_stats[hour]['total_rate'] += store.get('rate', 0)
            time_stats[hour]['working_staff'] += store.get('working_staff', 0)
            time_stats[hour]['active_staff'] += store.get('active_staff', 0)
        
        # データフレームに変換
        time_df = pd.DataFrame([
            {
                '時間帯': f'{hour}:00',
                '対象店舗数': stats['store_count'],
                '平均稼働率': round(stats['total_rate'] / stats['store_count'], 1) if stats['store_count'] > 0 else 0,
                '総勤務人数': stats['working_staff'],
                '総即ヒメ数': stats['active_staff'],
                '平均待機率': round(stats['active_staff'] / stats['working_staff'] * 100, 1) if stats['working_staff'] > 0 else 0
            }
            for hour, stats in sorted(time_stats.items())
        ])
        
        # 時間帯別シートに出力
        time_df.to_excel(writer, sheet_name='時間帯別分析', index=False)
        ws = writer.sheets['時間帯別分析']
        
        # グラフの追加
        chart = BarChart()
        chart.title = "時間帯別平均稼働率"
        chart.y_axis.title = '稼働率 (%)'
        chart.x_axis.title = '時間帯'
        
        data = Reference(ws, min_col=3, min_row=1, max_row=len(time_df)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(time_df)+1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 2
        
        ws.add_chart(chart, "H2")
        
    def _create_genre_analysis_sheet(self, writer, store_details):
        """ジャンル分析シートを作成"""
        df = pd.DataFrame(store_details)
        
        # ジャンル別の集計
        genre_stats = df.groupby(['業種', 'ジャンル']).agg({
            '店舗名': 'count',
            '稼働率': ['mean', 'min', 'max'],
            '勤務人数': 'sum',
            '即ヒメ数': 'sum'
        }).round(1).reset_index()
        
        # 列名を整理
        genre_stats.columns = [
            '業種',
            'ジャンル',
            '店舗数',
            '平均稼働率',
            '最小稼働率',
            '最大稼働率',
            '総勤務人数',
            '総即ヒメ数'
        ]
        
        # 稼働率でソート
        genre_stats = genre_stats.sort_values(['業種', '平均稼働率'], ascending=[True, False])
        
        # シートに出力
        genre_stats.to_excel(writer, sheet_name='ジャンル分析', index=False)
        ws = writer.sheets['ジャンル分析']

            return output_path
            
        except Exception as e:
            raise Exception(f"レポート生成中にエラーが発生しました: {str(e)}")

    def _apply_sheet_styling(self, ws, df):
        """シートの基本スタイル適用"""
        # ヘッダー行の高さを調整
        ws.row_dimensions[1].height = 25
        
        for col in range(1, len(df.columns) + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            # 列幅の自動調整（最小幅15、最大幅40）
            column_letter = cell.column_letter
            max_length = max([len(str(cell.value))] + 
                           [len(str(r[col-1])) if r[col-1] else 0 for r in df.values])
            adjusted_width = min(max(15, (max_length + 2)), 40)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = self.border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(name='Yu Gothic')

    def _apply_conditional_formatting(self, ws, df):
        """稼働率に応じた条件付き書式を適用"""
        rate_col = [i for i, col in enumerate(df.columns) if col == '稼働率'][0] + 1
        color_scale = ColorScaleRule(
            start_type='num', start_value=0, start_color='FF0000',
            mid_type='num', mid_value=50, mid_color='FFFF00',
            end_type='num', end_value=100, end_color='00FF00'
        )
        ws.conditional_formatting.add(f"{chr(64+rate_col)}2:{chr(64+rate_col)}{len(df)+1}", color_scale)

    def _create_area_analysis_sheet(self, writer, store_details):
        """エリア分析シートを作成"""
        df = pd.DataFrame(store_details)
        
        # エリア分析の集計
        area_stats = df.groupby('エリア').agg({
            '店舗名': 'count',
            '稼働率': ['mean', 'min', 'max'],
            '勤務人数': ['sum', 'mean'],
            '即ヒメ数': ['sum', 'mean']
        }).round(1).reset_index()
        
        # 列名を分かりやすく変更
        area_stats.columns = [
            'エリア',
            '店舗数',
            '平均稼働率',
            '最小稼働率',
            '最大稼働率',
            '総勤務人数',
            '平均勤務人数',
            '総即ヒメ数',
            '平均即ヒメ数'
        ]
        
        # 稼働率でソート
        area_stats = area_stats.sort_values('平均稼働率', ascending=False)
        
        # エリア分析シートに出力
        area_stats.to_excel(writer, sheet_name='エリア分析', index=False)
        ws = writer.sheets['エリア分析']
        
        # 条件付き書式の追加（稼働率の列に色付け）
        rate_columns = ['平均稼働率', '最小稼働率', '最大稼働率']
        for col_idx, col_name in enumerate(rate_columns, start=3):
            color_scale = ColorScaleRule(
                start_type='num', start_value=0, start_color='FF6B6B',
                mid_type='num', mid_value=50, mid_color='FFD93D',
                end_type='num', end_value=100, end_color='6BCB77'
            )
            ws.conditional_formatting.add(f"{chr(64+col_idx)}2:{chr(64+col_idx)}{len(area_stats)+1}", color_scale)
        
        # グラフの追加
        chart = BarChart()
        chart.title = "エリア別平均稼働率"
        chart.y_axis.title = '稼働率 (%)'
        chart.x_axis.title = 'エリア'
        
        data = Reference(ws, min_col=3, min_row=1, max_row=len(area_stats)+1)
        cats = Reference(ws, min_col=1, min_row=2, max_row=len(area_stats)+1)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.style = 2
        
        ws.add_chart(chart, "E2")

    def _adjust_column_widths(self, ws):
        """列幅の自動調整"""
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    def generate_store_report(self, store_data, output_path):
        pass  # 未実装の機能
